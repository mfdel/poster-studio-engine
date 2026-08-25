#!/usr/bin/env python3
"""Build the SKU management workbook for the Hopscotch Maps Etsy listing.

Produces studio/print_on_demand/sku_manager.xlsx with two sheets:
  - Assumptions : editable inputs (FX, VAT, Etsy fees). Change these and every
                  cost/margin below recalculates.
  - SKUs        : one row per fulfillable product variation, with our SKU, the
                  Prodigi SKU + attributes, retail price, Prodigi product/shipping
                  cost, Etsy fees, VAT, total cost, net profit and margin.

Prodigi costs are LIVE quotes (item + Budget shipping, to Sweden, ex-VAT) pulled
2026-07-31 via studio/commerce/prodigi_quotes.py. Re-run that script to refresh, then edit
the PRODIGI_* dicts below and rebuild:  uv run python studio/commerce/build_sku_workbook.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent.parent / "print_on_demand" / "sku_manager.xlsx"

# --- LIVE Prodigi costs to SE (EUR, ex-VAT): size -> (prodigi_sku, item, shipping) ---
POSTER = {  # printed poster tier — GLOBAL-BLP (Budget poster, Silk 170gsm); cm sizes only
    "30×40 cm": ("GLOBAL-BLP-12X16", 5.00, 5.40),
    "40×50 cm": ("GLOBAL-BLP-16X20", 5.00, 5.90),
    "50×70 cm": ("GLOBAL-BLP-20X28", 7.00, 7.00),
}
FRAMED = {  # framed tier — GLOBAL-BFP (Budget Framed Poster, no mount); cost equal across colors
    "30×40 cm": ("GLOBAL-BFP-12X16", 19.00, 13.95),
    "40×50 cm": ("GLOBAL-BFP-16X20", 24.00, 13.95),
    "50×70 cm": ("GLOBAL-BFP-20X28", 30.00, 19.40),
}
FRAME_COLORS = [("Wood", "natural"), ("White", "white")]  # budget frame: no black on rectangular

# Per-size retail prices (EUR). Size is now a PRICED Etsy variation, so each size
# has its own price (and its own single SKU). Edit freely.
PRICE = {
    "digital": 24,
    "poster": {"30×40 cm": 34, "40×50 cm": 39, "50×70 cm": 44},
    "framed": {"30×40 cm": 79, "40×50 cm": 89, "50×70 cm": 109},
}

# --- Region cost spread (LIVE Prodigi total, item+ship, to each country, EUR) ---
# 50×70 poster (BLP-20X28) and 50×70 framed (BFP-20X28) — the price-cap sizes.
REGION_COST = {  # country -> (poster_50x70, framed_50x70 or None if no fulfilment)
    "SE": (14.00, 49.40), "DE": (17.06, 42.95), "FR": (12.90, 46.15),
    "NL": (15.05, 41.85), "GB": (12.68, 47.28), "US": (16.14, 54.24),
    "CA": (24.32, 114.72), "AU": (22.04, 165.21), "JP": (24.66, 155.23),
}
BASELINE = "SE"
ETSY_SHIP_FEE = 0.065  # Etsy transaction fee also applies to shipping → gross up surcharges

# Recommended shipping-surcharge tiers (EUR), derived from the deltas above and
# grossed up for the Etsy shipping fee. framed_surcharge=None means "don't offer framed".
REGION_TIERS = [
    ("Europe (EU/EEA + UK)", "SE · DE · FR · NL · GB · IT · ES · …", 0, 0,
     "Costs within ±€3 of home (some cheaper) — flat price already covers it."),
    ("United States", "US", 0, 6,
     "Poster ≈ home; framed +~€5. NOTE: framed 30×40 is not fulfilled in the US — offer 40×50/50×70 only."),
    ("Canada", "CA", 13, 70,
     "Poster +~€12; framed +~€37–65 (frame would total ~€179). Consider poster + digital only."),
    ("Australia / Japan / Rest of world", "AU · JP · NZ · …", 12, None,
     "Poster +~€8–11 (OK). Framed +€78–116 — NOT viable; ship digital + poster only."),
]

# ------------------------------------------------------------------ styling
HEAD_FILL = PatternFill("solid", fgColor="4A3B2A")
SUB_FILL = PatternFill("solid", fgColor="EFE7D8")
DIG_FILL = PatternFill("solid", fgColor="EAF2E4")
POS_FILL = PatternFill("solid", fgColor="FBF6EC")
FRM_FILL = PatternFill("solid", fgColor="F5ECFB")
WHITE = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D8D2C4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
EUR = '#,##0.00" €"'
SEK = '#,##0" kr"'
PCT = '0.0%'


def build():
    wb = Workbook()

    # ============================================================ Assumptions
    a = wb.active
    a.title = "Assumptions"
    a["A1"] = "Assumptions — edit these; the SKUs sheet recalculates"
    a["A1"].font = Font(bold=True, size=13, color="4A3B2A")
    a.merge_cells("A1:C1")

    rows = [
        ("EUR → SEK rate", 11.30, "Update to the live rate before pricing in SEK"),
        ("Swedish VAT rate", 0.25, "Standard rate"),
        ("VAT registered?", "No", 'Yes / No. "No" = below SEK 120,000 threshold'),
        ("Etsy transaction fee", 0.065, "% of item price"),
        ("Etsy payment processing %", 0.04, "Sweden"),
        ("Etsy payment fixed (EUR)", 0.27, "~3 kr per order"),
        ("Etsy listing fee (EUR)", 0.18, "$0.20 per listing / renewal"),
        ("Offsite Ads fee", 0.00, "Set 0.15 only for orders that trigger an Offsite Ad"),
    ]
    labels = {}
    for i, (label, val, note) in enumerate(rows, start=2):
        a[f"A{i}"] = label
        a[f"A{i}"].font = BOLD
        a[f"B{i}"] = val
        a[f"C{i}"] = note
        a[f"C{i}"].font = Font(italic=True, color="9A8A72")
        labels[label] = f"'Assumptions'!$B${i}"
        if isinstance(val, float) and val < 1:
            a[f"B{i}"].number_format = PCT if "fee" in label or "rate" in label.lower() and "SEK" not in label else EUR
    a["B2"].number_format = '0.00'
    a["B7"].number_format = EUR
    a["B8"].number_format = EUR
    a["B3"].number_format = PCT
    a.column_dimensions["A"].width = 26
    a.column_dimensions["B"].width = 12
    a.column_dimensions["C"].width = 52

    FX, VATr, VATreg = labels["EUR → SEK rate"], labels["Swedish VAT rate"], labels["VAT registered?"]
    TXN = labels["Etsy transaction fee"]
    PAY = labels["Etsy payment processing %"]
    PAYFIX = labels["Etsy payment fixed (EUR)"]
    LIST = labels["Etsy listing fee (EUR)"]
    OFF = labels["Offsite Ads fee"]

    # ================================================================= SKUs
    s = wb.create_sheet("SKUs")
    headers = [
        ("Category", 13), ("Size", 12), ("Frame color", 11),
        ("SKU", 20), ("Prodigi SKU", 18),
        ("Prodigi attributes", 20), ("Retail (EUR)", 12),
        ("Prodigi product (EUR)", 13), ("Prodigi shipping (EUR)", 13),
        ("Prodigi total (EUR)", 13), ("Etsy fees (EUR)", 12),
        ("Input VAT (EUR)", 12), ("Output VAT (EUR)", 12),
        ("Total cost (EUR)", 13), ("Net profit (EUR)", 13),
        ("Margin %", 10), ("Retail (SEK)", 12), ("Notes", 34),
    ]
    for c, (name, width) in enumerate(headers, start=1):
        cell = s.cell(row=1, column=c, value=name)
        cell.fill = HEAD_FILL
        cell.font = WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        s.column_dimensions[get_column_letter(c)].width = width
    s.row_dimensions[1].height = 34

    # column letters (G=retail … R=notes)
    G, H, I, J, K, L, M, N, O, P, Q, R = "G H I J K L M N O P Q R".split()

    def add_row(r, cat, size, color, sku, prodigi_sku, attrs,
                retail, item, ship, fill, note=""):
        vals = [cat, size, color, sku, prodigi_sku, attrs, retail, item, ship]
        for ci, v in enumerate(vals, start=1):
            s.cell(row=r, column=ci, value=v)
        # formulas
        s[f"{J}{r}"] = f"={H}{r}+{I}{r}"                                   # prodigi total
        s[f"{K}{r}"] = f"={G}{r}*({TXN}+{PAY}+{OFF})+{PAYFIX}+{LIST}"       # etsy fees
        s[f"{L}{r}"] = f'=IF({VATreg}="No",({J}{r}+{K}{r})*{VATr},0)'       # irrecoverable input VAT
        s[f"{M}{r}"] = f'=IF({VATreg}="Yes",{G}{r}*{VATr}/(1+{VATr}),0)'    # output VAT on sale
        s[f"{N}{r}"] = f"={J}{r}+{K}{r}+{L}{r}+{M}{r}"                      # total cost
        s[f"{O}{r}"] = f"={G}{r}-{N}{r}"                                    # net profit
        s[f"{P}{r}"] = f"=IF({G}{r}=0,0,{O}{r}/{G}{r})"                     # margin
        s[f"{Q}{r}"] = f"={G}{r}*{FX}"                                      # retail SEK
        s[f"{R}{r}"] = note
        # formats + borders
        for ci in range(1, 19):
            cell = s.cell(row=r, column=ci)
            cell.border = BORDER
            cell.fill = fill
            if ci in (7, 8, 9, 10, 11, 12, 13, 14, 15):
                cell.number_format = EUR
                cell.alignment = CENTER
            elif ci == 16:
                cell.number_format = PCT
                cell.alignment = CENTER
            elif ci == 17:
                cell.number_format = SEK
                cell.alignment = CENTER
            elif ci in (1, 3):
                cell.alignment = CENTER
            elif ci == 18:
                cell.alignment = LEFT
        s[f"{O}{r}"].font = BOLD
        s[f"{P}{r}"].font = BOLD

    size_code = {"30×40 cm": "3040", "40×50 cm": "4050", "50×70 cm": "5070"}

    r = 2
    # --- Digital (one price, no size variation — all sizes live in the file) ---
    add_row(r, "Digital", "all sizes", "—", "PGMAP-DIG", "—",
            "delivered as PDF", PRICE["digital"], 0, 0, DIG_FILL,
            "One SKU, one price. All sizes included in the file — no size variation.")
    r += 1

    # --- Printed poster (price varies by size; one SKU per size) ---
    for size, (psku, item, ship) in POSTER.items():
        add_row(r, "Printed poster", size, "—", f"PGMAP-POS-{size_code[size]}",
                psku, "Budget Silk 170gsm", PRICE["poster"][size], item, ship,
                POS_FILL, "Size = priced Etsy variation.")
        r += 1

    # --- Framed poster (price varies by size; SKU per size × frame color) ---
    for size, (fsku, item, ship) in FRAMED.items():
        for cname, cval in FRAME_COLORS:
            add_row(r, "Framed poster", size, cname,
                    f"PGMAP-FRM-{cname[:2].upper()}-{size_code[size]}",
                    fsku, f"Budget, color={cval}", PRICE["framed"][size], item, ship,
                    FRM_FILL, "Size + frame color = priced Etsy variations. Cost equal across colors.")
            r += 1

    last = r - 1

    # totals / averages strip
    r += 1
    s.cell(row=r, column=1, value="Reference").font = BOLD
    s.cell(row=r, column=6, value="Poster avg net →").alignment = Alignment(horizontal="right")
    s.cell(row=r, column=15, value=f'=AVERAGEIF(A2:A{last},"Printed poster",{O}2:{O}{last})')
    s.cell(row=r, column=15).number_format = EUR
    s.cell(row=r, column=15).font = BOLD
    r += 1
    s.cell(row=r, column=6, value="Framed avg net →").alignment = Alignment(horizontal="right")
    s.cell(row=r, column=15, value=f'=AVERAGEIF(A2:A{last},"Framed poster",{O}2:{O}{last})')
    s.cell(row=r, column=15).number_format = EUR
    s.cell(row=r, column=15).font = BOLD

    # notes under the table
    r += 2
    for line in [
        "Prodigi costs = live quotes (item + Budget shipping) to Sweden, ex-VAT, 2026-07-31. Refresh with studio/commerce/prodigi_quotes.py.",
        "Costs rise for overseas ship-to (e.g. US 50×70 poster ≈ €16.1; US 50×70 framed ≈ €54.2). Model shows the SE home market — see the Region surcharges sheet.",
        "Prices vary by SIZE (a priced Etsy variation). One SKU per variation — the SKU is what you enter on Etsy AND use to pick the Prodigi product.",
        'Input VAT (irrecoverable) applies while "VAT registered?" = No. Register → set to Yes: output VAT applies and input VAT becomes reclaimable (set to 0 here).',
        "Frame color: Wood → Prodigi 'natural', White → 'white' (budget frame has no black on rectangular sizes).",
    ]:
        c = s.cell(row=r, column=1, value=line)
        c.font = Font(italic=True, color="9A8A72", size=9)
        s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=18)
        r += 1

    s.freeze_panes = "E2"
    s.sheet_view.showGridLines = False

    build_regions(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"saved {OUT}  ({last - 1} variation rows)")


def build_regions(wb):
    """Third sheet: recommended per-region shipping surcharges + the raw cost spread."""
    g = wb.create_sheet("Region surcharges")
    g["A1"] = "Region-based shipping surcharges"
    g["A1"].font = Font(bold=True, size=13, color="4A3B2A")
    g.merge_cells("A1:G1")
    g["A2"] = ("Etsy can't vary the ITEM price by country — only shipping. Set these as per-region "
               "shipping charges. Amounts are grossed up for Etsy's 6.5% shipping fee.")
    g["A2"].font = Font(italic=True, color="9A8A72", size=9)
    g.merge_cells("A2:G2")

    # ---- recommendation table ----
    heads = [("Region tier", 30), ("Example countries", 34), ("Poster surcharge", 15),
             ("Framed surcharge", 15), ("Poster total (50×70)", 14),
             ("Framed total (50×70)", 15), ("Notes", 60)]
    hr = 4
    for c, (name, w) in enumerate(heads, start=1):
        cell = g.cell(row=hr, column=c, value=name)
        cell.fill = HEAD_FILL
        cell.font = WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        g.column_dimensions[get_column_letter(c)].width = w
    g.row_dimensions[hr].height = 30

    poster_cap = max(PRICE["poster"].values())   # 50×70 price (surcharges are worst-case)
    framed_cap = max(PRICE["framed"].values())
    r = hr + 1
    for tier, examples, pos_s, frm_s, note in REGION_TIERS:
        g.cell(row=r, column=1, value=tier).font = BOLD
        g.cell(row=r, column=2, value=examples)
        g.cell(row=r, column=3, value=pos_s).number_format = EUR
        if frm_s is None:
            g.cell(row=r, column=4, value="— (don't ship)")
            g.cell(row=r, column=6, value="digital + poster only")
        else:
            g.cell(row=r, column=4, value=frm_s).number_format = EUR
            g.cell(row=r, column=6, value=f"={framed_cap}+D{r}").number_format = EUR
        g.cell(row=r, column=5, value=f"={poster_cap}+C{r}").number_format = EUR
        g.cell(row=r, column=7, value=note)
        for c in range(1, 8):
            cell = g.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c in (2, 7) else CENTER
        r += 1

    # ---- raw cost spread reference ----
    r += 1
    g.cell(row=r, column=1, value="Raw Prodigi cost by ship-to (50×70, item+ship, EUR) — the price-cap sizes").font = BOLD
    g.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    countries = list(REGION_COST.keys())
    g.cell(row=r, column=1, value="Product").font = BOLD
    for i, cc in enumerate(countries):
        cell = g.cell(row=r, column=2 + i, value=cc)
        cell.font = BOLD
        cell.alignment = CENTER
    r += 1
    base_frm = REGION_COST[BASELINE][1]
    for idx, prod in enumerate(("Poster 50×70", "Framed 50×70")):
        g.cell(row=r, column=1, value=prod)
        for i, cc in enumerate(countries):
            val = REGION_COST[cc][idx]
            cell = g.cell(row=r, column=2 + i, value=val if val is not None else "—")
            if isinstance(val, (int, float)):
                cell.number_format = EUR
            cell.alignment = CENTER
        r += 1
    # delta row for framed (the one that bites)
    g.cell(row=r, column=1, value="Framed Δ vs SE").font = Font(italic=True, color="9A8A72")
    for i, cc in enumerate(countries):
        val = REGION_COST[cc][1]
        d = (val - base_frm) if val is not None else None
        cell = g.cell(row=r, column=2 + i, value=(f"+{d:.0f}" if d is not None else "—"))
        cell.alignment = CENTER
        cell.font = Font(italic=True, color="9A8A72")
    r += 2

    for line in [
        "Surcharge = (region cost − home cost), grossed up for Etsy's 6.5% shipping fee, rounded up.",
        "Europe: costs sit within ±€3 of home (several cheaper) — no surcharge needed.",
        "Framed to Canada/Australia/Japan ships internationally (no local frame) → set framed shipping to EU/US only.",
        "Refresh the raw numbers with studio/commerce/prodigi_quotes.py (change SNAPSHOT_COUNTRIES) and edit REGION_COST here.",
    ]:
        c = g.cell(row=r, column=1, value=line)
        c.font = Font(italic=True, color="9A8A72", size=9)
        g.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1

    g.sheet_view.showGridLines = False


if __name__ == "__main__":
    build()
